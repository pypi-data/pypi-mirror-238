"""
Test case for BlendSequenceStartCell.
.. since: 0.1
"""

# -*- coding: utf-8 -*-
# Copyright (c) 2022 Endeavour Mining
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to read
# the Software only. Permissions is hereby NOT GRANTED to use, copy, modify,
# merge, publish, distribute, sublicense, and/or sell copies of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NON-INFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

from datetime import date
import openpyxl  # type: ignore
from hamcrest import assert_that, is_, equal_to
from edv_dwh_connector.blend_proposal.excel.blend_proposal_start_cell \
    import BlendProposalStartCell
from edv_dwh_connector.blend_proposal.excel.blend_sequence_start_cell \
    import BlendSequenceStartCell


BLEND_FILE_NAME = "3.Daily_Blend_Template_process_October.xlsx"


def test_finds_first_start_cell(resource_path_root) -> None:
    """
    Tests that it finds first start cell.
    :param resource_path_root: Resource path
    """
    sheet = openpyxl.load_workbook(
        resource_path_root / BLEND_FILE_NAME
    ).active
    sequence = BlendSequenceStartCell(
        sheet=sheet,
        bp_start_cell=BlendProposalStartCell(
            sheet=sheet,
            day=date(2022, 10, 7)
        ),
        number=1
    )
    assert_that(
        sequence.exist(), is_(True),
        "First sequence should exist"
    )
    assert_that(
        sequence.row(), equal_to(11),
        "First sequence start cell row should match"
    )
    assert_that(
        sequence.column(), equal_to(9),
        "First sequence start cell column should match"
    )


def test_finds_other_start_cell(resource_path_root) -> None:
    """
    Tests that it finds another start cell.
    :param resource_path_root: Resource path
    """
    sheet = openpyxl.load_workbook(
        resource_path_root / BLEND_FILE_NAME
    ).active
    sequence = BlendSequenceStartCell(
        sheet=sheet,
        bp_start_cell=BlendProposalStartCell(
            sheet=sheet,
            day=date(2022, 10, 7)
        ),
        number=2
    )
    assert_that(
        sequence.exist(), is_(True),
        "Other sequence should exist"
    )
    assert_that(
        sequence.row(), equal_to(27),
        "Other sequence start cell row should match"
    )
    assert_that(
        sequence.column(), equal_to(9),
        "Other sequence start cell column should match"
    )


def test_finds_not_start_cell(resource_path_root) -> None:
    """
    Tests that it doesn't find non-existent cell.
    :param resource_path_root: Resource path
    """
    sheet = openpyxl.load_workbook(
        resource_path_root / BLEND_FILE_NAME
    ).active
    sequence = BlendSequenceStartCell(
        sheet=sheet,
        bp_start_cell=BlendProposalStartCell(
            sheet=sheet,
            day=date(2022, 10, 7)
        ),
        number=5
    )
    assert_that(sequence.exist(), is_(False))
