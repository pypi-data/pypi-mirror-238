"""
This module defines Blend proposal start cell.
.. since: 0.1
"""

# -*- coding: utf-8 -*-
# Copyright (c) 2022 Endeavour Mining
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to read
# the Software only. Permissions is hereby NOT GRANTED to use, copy, modify,
# merge, publish, distribute, sublicense, and/or sell copies of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NON-INFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

from datetime import date
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore
from edv_dwh_connector.blend_proposal.excel.start_cell import StartCell


class BlendProposalStartCell(StartCell):
    """
    Blend proposal start cell.
    .. since: 0.1
    """

    def __init__(self, sheet: Worksheet, day: date):
        """
        Ctor.
        :param sheet: Excel sheet
        :param day: Day to find
        """
        super().__init__()
        self.__sheet = sheet
        self.__day = day

    def _find(self) -> tuple:
        frow, fcolumn = -1, -1
        for col in range(1, self.__sheet.max_column):
            for row in range(1, 100):
                current = self.__sheet.cell(row, col)
                after = self.__sheet.cell(row, col + 1)
                if current.value == 'Date' and after.is_date and \
                        after.value.date() == self.__day:
                    frow, fcolumn = row, col
                    break
            if frow != -1:
                break
        return frow, fcolumn
