"""
This module defines Excel blend sequence.
.. since 0.1
"""

# -*- coding: utf-8 -*-
# Copyright (c) 2022 Endeavour Mining
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to read
# the Software only. Permissions is hereby NOT GRANTED to use, copy, modify,
# merge, publish, distribute, sublicense, and/or sell copies of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NON-INFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

import re
from datetime import date

from openpyxl.worksheet.worksheet import Worksheet  # type: ignore

from edv_dwh_connector.blend_proposal.blend_material \
    import BlendMaterials
from edv_dwh_connector.blend_proposal.blend_sequence \
    import BlendSequence, BlendSequences
from edv_dwh_connector.blend_proposal.excel.blend_sequence_start_cell \
    import BlendSequenceStartCell
from edv_dwh_connector.blend_proposal.excel.sequence_total_row\
    import SequenceTotalRow
from edv_dwh_connector.blend_proposal.excel.excel_blend_material \
    import ExcelBlendMaterials
from edv_dwh_connector.blend_proposal.excel.start_cell import StartCell
from edv_dwh_connector.exceptions import ValueNotFoundError


class ExcelBlendSequence(BlendSequence):
    """
    Excel blend sequence.
    .. since: 0.1
    """

    def __init__(
        self, sheet: Worksheet, order: int, bp_start_cell: StartCell,
        bs_start_cell: StartCell
    ):
        """
        Ctor.
        :param sheet: Worksheet
        :param order: Sequence order
        :param bp_start_cell: Blend proposal start cell
        :param bs_start_cell: Blend sequence start cell
        """
        self.__sheet = sheet
        self.__order = order
        self.__bp_start_cell = bp_start_cell
        self.__bs_start_cell = bs_start_cell

    def order(self) -> int:
        return self.__order

    def date(self) -> date:
        return self.__sheet.cell(
            self.__bp_start_cell.row(), self.__bp_start_cell.column() + 1
        ).value.date()

    def name(self) -> str:
        return self.__sheet.cell(
            self.__bs_start_cell.row(), self.__bs_start_cell.column() + 1
        ).value.strip().strip('_')

    def average_au_grade(self) -> float:
        return self.__value_of("average")

    def soluble_copper(self) -> float:
        return self.__value_of("soluble")

    def arsenic(self) -> float:
        try:
            value = self.__value_of("as")
        except ValueNotFoundError:
            value = self.__value_of("arsenic")
        return value

    def moisture_estimated(self) -> float:
        return self.__value_of("moisture")

    def oxide_transition(self) -> float:
        return self.__value_of("oxide/transition")

    def fresh(self) -> float:
        return self.__value_of("fresh")

    def recovery_blend_estimated(self) -> float:
        return self.__value_of("recovery")

    def materials(self) -> BlendMaterials:
        return ExcelBlendMaterials(self.__sheet, self.__bs_start_cell)

    def __value_of(self, word) -> float:
        """
        Gets value of key containing word.
        :param word: Word
        :return: Value
        :raises ValueNotFoundError: If value not found
        """
        value = None
        rrow = self.__bs_start_cell.row()
        rcol = self.__bs_start_cell.column()
        count = SequenceTotalRow(self.__sheet, self.__bs_start_cell).count()
        for row in range(rrow, rrow + count):
            current = None
            for i in range(1, 5):
                if self.__sheet.cell(row, rcol - i).value is not None:
                    current = self.__sheet.cell(row, rcol - i)
                    break
            if current is None:
                continue
            if current.value is not None and \
                    re.search(word, current.value, re.IGNORECASE):
                for i in range(1, 10):
                    val = self.__sheet.cell(row, rcol + i).value
                    if val is not None:
                        if str(val).replace('.', '', 1).isdigit():
                            value = val
                        else:
                            value = float('NaN')
                        break
                if value is None:
                    value = float('NaN')
            if value is not None:
                break
        if value is None:
            raise ValueNotFoundError(
                f'Unable to find blend characteristic {word}'
            )
        return value


class ExcelBlendSequences(BlendSequences):
    """
    Excel blend sequences.
    .. since: 0.5
    """

    def __init__(self, sheet: Worksheet, bp_start_cell: StartCell) -> None:
        """
        Ctor.
        :param sheet: Worksheet
        :param bp_start_cell: Blend proposal start cell
        """
        self.__sheet = sheet
        self.__bp_start_cell = bp_start_cell

    def items(self) -> list:
        items = []
        for number in range(1, 10):
            try:
                scell = BlendSequenceStartCell(
                    self.__sheet, self.__bp_start_cell, number
                )
                if scell.exist():
                    items.append(
                        ExcelBlendSequence(
                            self.__sheet, number, self.__bp_start_cell, scell
                        )
                    )
            except ValueError:
                break
        return items

    def add(self, sequence: BlendSequence) -> BlendSequence:
        raise NotImplementedError(
            "We don't support adding sequence to an Excel file"
        )
