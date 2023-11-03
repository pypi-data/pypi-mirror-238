"""
This module defines Excel sequence material.
.. since: 0.1
"""

# -*- coding: utf-8 -*-
# Copyright (c) 2022 Endeavour Mining
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to read
# the Software only. Permissions is hereby NOT GRANTED to use, copy, modify,
# merge, publish, distribute, sublicense, and/or sell copies of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NON-INFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

import re

from openpyxl.cell import Cell  # type: ignore
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore

from edv_dwh_connector.blend_proposal.blend_material import BlendMaterial
from edv_dwh_connector.blend_proposal.blend_material \
    import BlendMaterials
from edv_dwh_connector.blend_proposal.excel.sequence_total_row \
    import SequenceTotalRow
from edv_dwh_connector.blend_proposal.excel.blend_material_start_cell \
    import BlendMaterialStartCell
from edv_dwh_connector.blend_proposal.excel.start_cell import StartCell
from edv_dwh_connector.utils.name\
    import NameInLowercaseWithoutSpecialCharacterAndUnit
from edv_dwh_connector.exceptions import ValueNotFoundError


class ExcelBlendMaterial(BlendMaterial):
    """
    Excel sequence material.
    .. since: 0.1
    """

    def __init__(
        self, sheet: Worksheet, bs_start_cell: StartCell,
        sm_start_cell: StartCell
    ) -> None:
        """
        Ctor.
        :param sheet: Worksheet
        :param bs_start_cell: Blend sequence start cell
        :param sm_start_cell: Sequence material start cell
        """

        self.__sheet = sheet
        self.__bs_start_cell = bs_start_cell
        self.__sm_start_cell = sm_start_cell

    def machine(self) -> str:
        if re.search("SURGE_BIN", self.name(), re.IGNORECASE) or \
                re.search("COS", self.name(), re.IGNORECASE):
            value = "SURGE BIN"
        else:
            value = "CRUSHER"
        return value

    def pit(self) -> str:
        return self.__sheet.cell(
            self.__sm_start_cell.row(), self.__pit_column()
        ).value.strip().strip('_')\
            .strip("*").replace(" ", "")

    def name(self) -> str:
        return self.__value_of("Material").value\
            .strip().strip('_')\
            .strip("*")

    def au_grade(self) -> float:
        return self.__value_number_of("Au_Grade (g/t)")

    def soluble_copper(self) -> float:
        return self.__value_number_of("Sol_Cu (ppm)")

    def arsenic(self) -> float:
        return self.__value_number_of("As_(ppm)")

    def moisture(self) -> float:
        return self.__value_number_of("Moisture (%)")

    def indicative_rec(self) -> float:
        return self.__value_number_of("Indicative Rec (%)")

    def bucket(self) -> float:
        return self.__value_number_of("Bucket")

    def available_tons(self) -> float:
        return self.__value_number_of("Available Tons")

    def proportion(self) -> float:
        return self.__value_number_of("Prop (%)")

    def __pit_column(self) -> int:
        """
        Gets PIT column.
        :return: Column
        """
        return self.__bs_start_cell.column() - 1

    def __value_of(self, name: str) -> Cell:
        """
        Gets value at position.
        :param name: Name
        :return: Cell
        :raises ValueNotFoundError: If name not found
        """
        col = -1
        for i in range(14):
            colname = self.__sheet.cell(
                self.__bs_start_cell.row() + 1,
                self.__sm_start_cell.column() + i
            ).value
            if colname is None:
                continue
            if NameInLowercaseWithoutSpecialCharacterAndUnit(
                colname
            ).value() == NameInLowercaseWithoutSpecialCharacterAndUnit(
                name
            ).value():
                col = i
                break
        if col == -1:
            raise ValueNotFoundError(
                f'Blend material: column for {name} not found'
            )
        return self.__sheet.cell(
            self.__sm_start_cell.row(), self.__sm_start_cell.column() + col
        )

    def __value_number_of(self, name: str) -> float:
        """
        Gets value number at position.
        :param name: Name
        :return: Value
        """
        value = self.__value_of(name).value
        if value is None or not str(value).replace('.', '', 1).isdigit():
            value = float('NaN')
        return value


class ExcelBlendMaterials(BlendMaterials):
    """
    Excel blend material.
    .. since: 0.5
    """

    def __init__(self, sheet: Worksheet, bs_start_cell: StartCell) -> None:
        """
        Ctor.
        :param sheet: Worksheet
        :param bs_start_cell: Blend material start cell
        """
        self.__sheet = sheet
        self.__bs_start_cell = bs_start_cell

    def items(self) -> list:
        items = []
        start = self.__bs_start_cell.row() + 1
        for row in range(
            start,
            start + SequenceTotalRow(
                self.__sheet, self.__bs_start_cell
            ).count()
        ):
            col = self.__bs_start_cell.column()
            name = self.__sheet.cell(row, col).value
            if isinstance(name, str) and name is not None and \
                    not re.search("Material", name, re.IGNORECASE):
                items.append(
                    ExcelBlendMaterial(
                        self.__sheet, self.__bs_start_cell,
                        BlendMaterialStartCell(row=row, column=col)
                    )
                )
        return items

    def add(self, material: BlendMaterial) -> BlendMaterial:
        raise NotImplementedError(
            "We don't support adding material to an Excel file"
        )
