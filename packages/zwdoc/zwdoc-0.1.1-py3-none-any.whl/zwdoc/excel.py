from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl import load_workbook

class Excel(object):
    """Excel Class

    .. code-block:: Python

        wb = Excel('my.xlsx')
        wb.sheetnames
        ws = wb[0]
        ws = wb['sheetA']
        ws['A4']
        cell_range = ws['A1':'C2']
        coll_range = ws['C:D']
        rows_range = ws[5:10]
        tuple(ws.rows), tuple(ws.columns)

    """
    path = property(lambda o: o.pth, lambda o, v: setattr(o, 'pth', v))
    active_sheet = property(lambda o: o.wb.active)
    sheet_count = property(lambda o: len(o.wb._sheets))
    sheets = property(lambda o: [Sheet(a) for a in o.wb.worksheets])

    def __init__(self, pth=None) -> None:
        self.pth = Path(pth) if pth else None
        self.ext = Path(pth).suffix if pth else None
        self.wb = None
    
    def __enter__(self) -> Excel:
        self.open()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self.save()
    
    def open(self) -> Excel:
        if self.wb:
            return self
        if self.pth:
            self.wb = load_workbook(str(self.pth))
        else:
            self.wb = Workbook()
        return self

    def save(self) -> Excel:
        if self.pth is None:
            return self
        self.pth.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(self.pth)
        return self
    
    def create_sheet(self, title = None, index = None) -> None:
        wb = self.wb
        title = title or f'New Sheet {len(wb._sheets)+1}'
        if index is None:
            wb.create_sheet(title)
        else:
            wb.create_sheet(title, index)
    
    def __getitem__(self, key):
        wb = self.wb
        if isinstance(key, int):
            idx = key
            names = wb.sheetnames
            if idx > len(names) - 1:
                raise KeyError('Worksheet {0} does not exist.'.format(key))
            return Sheet(wb[names[idx]])
        return Sheet(wb[key])
    
    def __getattr__(self, name):
        return getattr(self.wb, name)

class Sheet(object):
    def __init__(self, o) -> None:
        self._o = o

    def __getattr__(self, name):
        return getattr(self._o, name)